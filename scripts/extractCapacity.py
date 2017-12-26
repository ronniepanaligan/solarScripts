from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

#Could probably automate this even more
#For now 1st workbook is the capacity data workbook, change name to specific year
#For now 2nd workbook is the capacity data layout that we will be copying, difference is that capacityData2 uses 'Energy Source 1' and other uses 'technology', the 2014-2016 sheets are the only oens with the 'technology' column
#Save to 'capacity_' + year chosen

workbook = load_workbook('./sheets/capacitySheets/2005.xlsx') #Worksheet to read data from
first_sheet = workbook.get_sheet_names()[0]
generatorData = workbook.get_sheet_by_name(first_sheet)

wb = load_workbook('./sheets/capacityData2.xlsx') #worksheet to copy layout
sheet = wb.get_sheet_names()[0]
capacityData = wb.get_sheet_by_name(sheet)

nameplateFromSheet = column_index_from_string('L') #Nameplate Capacity
techOrSource = column_index_from_string('AA') #Energy Source 1
stateFromSheet = column_index_from_string('C') #State

for generatorDataRow in range(3, generatorData.max_row): #iterate through 2016 capcacity sheet, this is where the capacity is located
    if generatorData.cell(row=generatorDataRow, column=nameplateFromSheet).value != ' ': #check to see if Nameplate value is empty
        for stateCol in range(2, 52): #look for the state of the current row in the capacityData sheet
            if generatorData.cell(row=generatorDataRow, column=stateFromSheet).value == capacityData.cell(row=1, column=stateCol).value: #If the state of the current 2016 capacity row is found
                for capacityDataRow in range(3, 42): #iterate through the capacityData sheet to find correspoinding Technology
                    if generatorData.cell(row=generatorDataRow, column=techOrSource).value == capacityData.cell(row=capacityDataRow, column=column_index_from_string('A')).value: #if the technology is found write to the cell
                        capacityData.cell(row=capacityDataRow, column=stateCol).value += generatorData.cell(row=generatorDataRow, column=nameplateFromSheet).value
                        break
                    if capacityData == capacityData.max_row:
                        capacityData.cell(row=3, column=stateCol).value += generatorData.cell(row=generatorDataRow, column=nameplateFromSheet).value
wb.save(filename = './sheets/capacityByState/capacity_2005.xlsx') #save to this worksheet
