from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
import matplotlib.pyplot as plt
import numpy

wkbk = load_workbook('FILENAME') # workbook that holds data you want to plot
sheet = wkbk.get_sheet_names()[1]
data = wkbk.get_sheet_by_name(sheet)

dataset = []

# example usage using N column
for x in range(1, data.max_row+1):
    if data.cell(row=x, column=column_index_from_string('N')).value:
        dataset.append(data.cell(row=x, column=column_index_from_string('N')).value)

plt.hist(dataset, bins='auto')
plt.title("Histogram")
plt.show()
