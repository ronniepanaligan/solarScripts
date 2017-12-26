from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import csv

workbook = load_workbook('sheet1.xlsx') #Worksheet to read from
first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)

with open('newSheet1.csv', 'w', newline='') as csvfile: #csv file to write to
    filewriter = csv.writer(csvfile, delimiter=',', quoting=csv.QUOTE_MINIMAL)
    for row in range(6, worksheet.max_row):
        lrow = []
        lrow.append(worksheet.cell(row=row, column=column_index_from_string('A')).value)
        for row2 in range(column_index_from_string('CB'), column_index_from_string('CN')):
            lrow.append(worksheet.cell(row=row, column=row2).value)
        filewriter.writerow(lrow)
    print("done")
    csvfile.close()
