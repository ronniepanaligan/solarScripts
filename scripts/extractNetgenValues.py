from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import os

workbook = load_workbook('./sheets/data.xlsx')
first_sheet = workbook.get_sheet_names()[0]
template = workbook.get_sheet_by_name(first_sheet) #template that will be used to write data to

arr = os.listdir('./sheets/netgenSheets')

for x in arr:
    wkbk = load_workbook('./sheets/netgenSheets/'+x)
    sheet = wkbk.get_sheet_names()[0]
    worksheet = wkbk.get_sheet_by_name(sheet) #sheet that contains the data that needs to be extracted
    year = worksheet.cell(row=15, column=column_index_from_string('CS')).value
    counted = 0
    added = 0

    for worksheet_row in range(6, worksheet.max_row+1): #iterate through each row to find solar generators
        if worksheet.cell(row=worksheet_row, column=column_index_from_string('O')).value == 'SUN': #found a solar generator
            counted = counted + 1
            found = 0
            for template_col in range(0, 12): #find the correspoindingyear by iterating through template row1 columns
                if template.cell(row=1, column=template_col*12+4).value == year: #found the correspoinding year
                    lrow = [] #create a new list to hold the 12 values(Jan-Dec)
                    lrow.append(worksheet.cell(row=worksheet_row, column=column_index_from_string('A')).value) #save the plantID
                    for generator_values in range(column_index_from_string('CB'), column_index_from_string('CN')): #save values found from these columns
                        if worksheet.cell(row=worksheet_row, column=generator_values).value == '.':
                            lrow.append('NULL')
                        else:
                            lrow.append(worksheet.cell(row=worksheet_row, column=generator_values).value)
                    for template_ID in range(1, template.max_row+1): #search for plant ID in the template
                        if template.cell(row=template_ID, column=column_index_from_string('A')).value == lrow[0]: #found the plant ID
                            added = added + 1
                            found = 1
                            for row in range(0, 12): #copy the 12 values stored in lrow
                                template.cell(row=template_ID, column=template_col*12+4+row).value = lrow[row+1]
                            break
                    break
            if found == 0:
                print('Couldnt find plant ID: ', worksheet.cell(row=worksheet_row, column=column_index_from_string('A')).value)
    print('counted rows for year ', year, ' is: ', counted)
    print('added rows for year ', year, ' is: ', added)

workbook.save(filename = './sheets/ex.xlsx') #save to this worksheet
