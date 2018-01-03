from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string

workbook = load_workbook('./sheets/netgenSheets/2015.xlsx') #Worksheet to read from
first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)

wb = Workbook()
ws = wb.active

for row in range(1, worksheet.max_row+1):
    if worksheet.cell(row=row, column=column_index_from_string('O')).value == 'SUN':
        lrow = []
        lrow.append(worksheet.cell(row=row, column=column_index_from_string('A')).value)
        for col in range(column_index_from_string('CB'), column_index_from_string('CN')):
            if worksheet.cell(row=row, column=col).value == '.':
                lrow.append('NULL')
            else:
                lrow.append(worksheet.cell(row=row, column=col).value)
        ws.append(lrow)
wb.save(filename = './sheets/ex2.xlsx')
