from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
import numpy

# COPY CAPACITY FACTORS FROM SHEET 0 TO SHEET 1 AND CALCULATE THE SLOPE

wkbk = load_workbook('FILENAME')
first = wkbk.get_sheet_names()[0]
second = wkbk.get_sheet_names()[1]
data = wkbk.get_sheet_by_name(first) # WHERE THE CAPACITY FACTORS ARE
capacity = wkbk.get_sheet_by_name(second) # WHERE WE WANT TO COPY THE FACTORS TO AND FIND SLOPE

for data_row in range(4, data.max_row+1): # iterate through worksheet
    y = []
    x = []
    for data_col in range(0, 12): # FIRST GET ALL CAPACITY FACTORS FROM ROW data_row
        if data.cell(row=data_row, column=data_col*14+16).value:
            if data.cell(row=data_row, column=data_col*14+16).value > 0:
                y.append(data.cell(row=data_row, column=data_col*14+16).value)
    for val in range(0, len(y)):
        x.append(val)
    for d in x: # COPY THE CAPACITY FACTORS TO THE OTHER SHEET
        capacity.cell(row=data_row-2, column=3+d).value = y[d]
    if(len(x) > 0 and len(y) > 0): # SOLVE FOR SLOPE
        res = numpy.polyfit(x,y,1)
        capacity.cell(row=data_row-2, column=14).value = res[0]
wkbk.save(filename = 'FILENAME')
