# implemented on 3/26
# purpose is to extract nameplate and store by state and tech
# set flag to sort by state or by tech

from openpyxl import load_workbook, Workbook # standard lib I use to read excel workbooks
from openpyxl.utils import column_index_from_string # tool to read column by letters
import os # lib for reading user input

print("Sheets should be placed in a folder called sheets")
book = input("Please enter the name of the workbook including .xlsx: ")

wkbk = load_workbook('./sheets/'+ book) # workbook to load from
first = wkbk.get_sheet_names()[0] # sheet to read from
worksheet = wkbk.get_sheet_by_name(first) # template that will be used to write data to

print("Enter 0 for fuel type then state or")
var = int(input("Enter 1 for state then fuel type: ")) #ensure type of var is int

#If user input is incorrect, defualt to 0
if var != 1 and var != 0:
    print("invalid input, defaulting to 0")
    var = 0

State = ['AK', 'AL', 'AR', 'AZ', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA', 'HI', 'IA', 'ID', 'IN', 'IL', 'KS', 'KY', 'LA', 'ME', 'MD', 'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 'NH', 'NJ', 'NM', 'NY', 'NC', 'ND', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX', 'UT', 'VT', 'VA', 'WA', 'WV', 'WI', 'WY'];
#fuelType = ['Coal', 'Oil', 'Natural gas', 'Nuclear', 'Hydro', 'Biomass', 'Wind', 'Solar', 'Geothermal', 'Storage'];
fuel = ['SUN', 'COL', 'DFO', 'GEO', 'HPS', 'HYC', 'MLG', 'NG', 'NUC', 'OOG', 'ORW', 'OTH', 'PC', 'RFO', 'WND', 'WOC', 'WOO', 'WWW'];

# clear cells
for row in range(2, worksheet.max_row+1):
    for cell in range(1, worksheet.max_column+1):
    #for cell in range(1, 3):
        worksheet.cell(row=row, column=cell).value = None

# Add data to first to columns
if var == 0:
    row = 2
    for tech in fuel:
        for state in State:
            worksheet.cell(row=row, column=1).value = state
            worksheet.cell(row=row, column=2).value = tech
            row = row + 1
else:
    row = 2
    for state in State:
        for tech in fuel:
            worksheet.cell(row=row, column=1).value = state
            worksheet.cell(row=row, column=2).value = tech
            row = row + 1

year = 2007 # starting year
yearCol = 3 # starting column

# ask if user wants program to automate the process
automate = int(input("Enter 1 if you would like the program to attempt to automatically find the following columns: Fuel, State and netgen other wise 0 to input manually: "))
if automate != 1 and automate != 0:
    print("invalid input, defaulting to 0")
    automate = 0

arr = os.listdir('./sheets/netgenSheets') # GET LIST OF SHEETS IN THIS DIRECTORY AND STORE IN LIST

for x in arr: #ITTERATE THROUGH NETGEN SHEETS
    w = load_workbook('./sheets/netgenSheets/'+x) # LOAD THIS WORKBOOK
    sheet = w.get_sheet_names()[0]
    data = w.get_sheet_by_name(sheet) #sheet that contains the data that needs to be extracted
    fuelColumn = 0 #Column that contains fuel code
    stateCol = 0 #Column that contains generators state
    netgenCol = 0 #Column that contains the netgen for the year

    if automate == 1:
        # attempt to find above columns
        for row in range(5, 9): #usually appears in these rows
            for col in range(1, data.max_column+1):
                val = data.cell(row=row, column=col).value # save cell value to this
                if type(val) == str: #print cell value and its col number
                    print(val, col)
                if val != None and type(val) == str: #Sometimes this is None and we cant subscript it so we will check if it is != None
                    if val[:3] == 'AER': #Check for AER Fuel code column
                        fuelColumn = col
                    if val[:14] == 'NET GENERATION' or val[:14] == 'Net Generation': #check for netgen total column
                        netgenCol = col
                if val == 'State' or val == 'Plant State': #find state col
                    stateCol = col

    #If any of the above columns == 0, then program couldnt find the correct columns so user must manually input the correct columns
    if fuelColumn == 0:
        print("Couldn't find column for fuel type")
        fuelColumn = int(input("Enter the column for the fuel type: "))
        # check here if user input is a number and in correct range(1, data.max_column+1)
        while fuelColumn < 0 or fuelColumn > data.max_column+1:
            print("Invalid input, please select number from 1 -", data.max_column)
            fuelColumn = int(input("Enter the column for the fuel type: "))
    if stateCol == 0:
        print("Couldn't find column for state")
        stateCol = int(input("Enter the column for generator's state: "))
        # check here if user input is a number and in correct range(1, data.max_column+1)
    if netgenCol == 0:
        print("Couldn't find column for netgen")
        netgenCol = int(input("Enter the column for the total net generation: "))
        # check here if user input is a number and in correct range(1, data.max_column+1)

    # extract netgen totals to correct cell
    for data_row in range(1, data.max_row+1): #first iterate through each row in data worksheet
        for r in range(1, worksheet.max_row+1): #Then iterate through main worksheet and find correspoinding fuel code and state
            if data.cell(row=data_row, column=fuelColumn).value == worksheet.cell(row=r, column=2).value and data.cell(row=data_row, column=stateCol).value == worksheet.cell(row=r, column=1).value:
                #found correct row to extract to
                if worksheet.cell(row=r, column=yearCol).value == None: #If the cell == None we need to make the cells value = 0 so that we can do some integer addition
                    worksheet.cell(row=r, column=yearCol).value = 0
                worksheet.cell(row=r, column=yearCol).value += data.cell(row=data_row, column=netgenCol).value # add to the cell

    print("Finished with sheet", year)
    year = year + 1
    yearCol = yearCol + 1

wkbk.save(filename = './sheets/newsheet.xlsx') #save to this worksheet
