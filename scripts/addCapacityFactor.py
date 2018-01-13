from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string

#add the capacity factor to main sheet

wkbk = load_workbook('./sheets/new1.xlsx')
first = wkbk.get_sheet_names()[0]
worksheet = wkbk.get_sheet_by_name(first) #template that will be used to write data to

for worksheet_row in range(4, worksheet.max_row+1): #iterate through worksheet
    print(worksheet_row)
    for worksheet_col in range(0, 12): #iterate through columns
        print('col')
        print(worksheet_col)
        if worksheet.cell(row=worksheet_row, column=worksheet_col*14+15).value and worksheet.cell(row=worksheet_row, column=worksheet_col*14+15).value != 0:
            total = 0
            for netgen_col in range(1, 13):
                if worksheet.cell(row=worksheet_row, column=worksheet_col*14+15-netgen_col).value and worksheet.cell(row=worksheet_row, column=worksheet_col*14+15-netgen_col).value != 'NULL':
                    total += worksheet.cell(row=worksheet_row, column=worksheet_col*14+15-netgen_col).value
                print(total)
            capacity = worksheet.cell(row=worksheet_row, column=worksheet_col*14+15).value
            print(capacity)
            val = total/(capacity*365*24)
            print(val)
            worksheet.cell(row=worksheet_row, column=worksheet_col*14+16).value = val
    print('added')
wkbk.save(filename = './sheets/cap_fac.xlsx')
