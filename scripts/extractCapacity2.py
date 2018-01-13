from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string

#extract nameplate capacity to main data sheet

wkbk = load_workbook('./sheets/net.xlsx')
first = wkbk.get_sheet_names()[0]
template = wkbk.get_sheet_by_name(first) #template that will be used to write data to

workbook = load_workbook('./sheets/capacitySheets/2016.xlsx')
first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)

energy_source = column_index_from_string('AH')
nameplate = column_index_from_string('P')
plantID = column_index_from_string('C')
year = 2016

for worksheet_row in range(1, worksheet.max_row+1):
    if worksheet.cell(row=worksheet_row, column=energy_source).value == 'SUN': #energy source
        for template_row in range(1, template.max_row+1):
            if template.cell(row=template_row, column=column_index_from_string('A')).value == worksheet.cell(row=worksheet_row, column=plantID).value:
                for template_col in range(0, 12):
                    if template.cell(row=1, column=template_col*14+15).value == year:
                        if template.cell(row=template_row, column=template_col*14+15).value:
                            template.cell(row=template_row, column=template_col*14+15).value += worksheet.cell(row=worksheet_row, column=nameplate).value
                        else:
                            template.cell(row=template_row, column=template_col*14+15).value = worksheet.cell(row=worksheet_row, column=nameplate).value
                        break
                break
wkbk.save(filename = './sheets/new1.xlsx')
